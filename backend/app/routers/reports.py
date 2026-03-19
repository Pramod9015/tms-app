"""
Reports router — export transactions as Excel, CSV, or PDF.
"""
import io
import csv
from datetime import date, datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.database import get_db
from app.models.transaction import Transaction
from app.models.bank import Bank
from app.models.user import User
from app.security.encryption import decrypt_field
from app.dependencies import get_current_user

router = APIRouter(prefix="/api/reports", tags=["Reports"])


async def _fetch_transactions(
    db: AsyncSession, user_id: int,
    date_from: Optional[date], date_to: Optional[date],
    bank_id: Optional[int], app_used: Optional[str],
):
    q = select(Transaction).where(Transaction.user_id == user_id)
    if date_from:
        q = q.where(Transaction.transaction_date >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(Transaction.transaction_date <= datetime.combine(date_to, datetime.max.time()))
    if bank_id:
        q = q.where(Transaction.bank_id == bank_id)
    if app_used:
        q = q.where(Transaction.app_used == app_used)
    q = q.order_by(Transaction.transaction_date.desc())
    result = await db.execute(q)
    return result.scalars().all()


async def _build_rows(db, transactions):
    rows = []
    for t in transactions:
        bank_name = ""
        if t.bank_id:
            r = await db.execute(select(Bank).where(Bank.id == t.bank_id))
            b = r.scalar_one_or_none()
            bank_name = b.bank_name if b else ""
        rows.append({
            "ID": t.id,
            "Date": t.transaction_date.strftime("%Y-%m-%d %H:%M"),
            "Type": t.transaction_type.value,
            "App": t.app_used.value,
            "Amount": float(t.amount),
            "Bank": bank_name,
            "Status": t.status.value,
            "Reference": t.reference_number or "",
            "Notes": decrypt_field(t.notes_enc) or "",
        })
    return rows


@router.get("/export/csv")
async def export_csv(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    bank_id: Optional[int] = Query(None),
    app_used: Optional[str] = Query(None),
):
    transactions = await _fetch_transactions(db, current_user.id, date_from, date_to, bank_id, app_used)
    rows = await _build_rows(db, transactions)

    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"},
    )


@router.get("/export/excel")
async def export_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    bank_id: Optional[int] = Query(None),
    app_used: Optional[str] = Query(None),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    transactions = await _fetch_transactions(db, current_user.id, date_from, date_to, bank_id, app_used)
    rows = await _build_rows(db, transactions)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    if rows:
        headers = list(rows[0].keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[key])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions.xlsx"},
    )


@router.get("/export/pdf")
async def export_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    bank_id: Optional[int] = Query(None),
    app_used: Optional[str] = Query(None),
):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    transactions = await _fetch_transactions(db, current_user.id, date_from, date_to, bank_id, app_used)
    rows = await _build_rows(db, transactions)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Transaction Report", styles["Title"]), Spacer(1, 12)]

    if rows:
        headers = list(rows[0].keys())
        table_data = [headers] + [[str(r[k]) for k in headers] for r in rows]
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
        ]))
        elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=transactions.pdf"},
    )
